import csv
import io
import os
from datetime import date, datetime, time, timedelta, timezone
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError
from sqlalchemy.orm import Session, joinedload
from app.db.session import get_db
from app.models.order import Order, OrderItem
from app.models.product import Product
from app.models.wallet import WalletTransaction
from app.models.worker import Worker
from app.schemas.admin import (
    ProductAdminOut,
    ProductCreateRequest,
    ProductUpdateRequest,
    OrderAdminOut,
    OrderEditRequest,
    OrderItemAdminOut,
    OrderRefundResult,
    WorkerOut,
    WorkerCreateRequest,
    WorkerUpdateRequest,
    WorkerSpending,
    WalletTopUpResult,
    WalletReversalResult,
)
from app.schemas.wallet import WalletTopUpRequest, WalletAdjustRequest, WalletTransactionOut
from app.core.security import hash_pin
from app.core.deps import require_admin
from app.core.product_categories import category_zh_for
from app.services.cashier_service import normalize_occurred_at
from app.services.storage import save_product_image, ALLOWED_IMAGE_EXTENSIONS

# A product at or below this many units on hand is flagged "low stock" across
# the app (admin dashboard alert, inventory chips, worker shop badge).
LOW_STOCK_THRESHOLD = 30
MAX_IMAGE_BYTES = 5 * 1024 * 1024
router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Products ─────────────────────────────────────────────────────────────────

@router.get("/products/", response_model=list[ProductAdminOut])
def admin_list_products(
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    products = db.query(Product).order_by(Product.name).all()
    result = []
    for p in products:
        out = ProductAdminOut.model_validate(p)
        out.low_stock = p.stock <= LOW_STOCK_THRESHOLD
        result.append(out)
    return result


@router.get("/products/export.xlsx")
def admin_export_products_xlsx(
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    """Whole catalogue as an .xlsx workbook, for stock-taking and offline
    review. One row per product; 'Stock value' = price x stock."""
    products = db.query(Product).order_by(Product.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "SKU", "Name", "Name (中文)", "Category", "Category (中文)",
        "Sub-category", "Brand", "Size", "Unit",
        "Price (Rp)", "Stock", "Stock value (Rp)", "Active",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for p in products:
        price = float(p.price)
        ws.append([
            p.sku, p.name, p.name_zh or "",
            p.category or "", p.category_zh or "",
            p.sub_category or "", p.brand or "", p.size or "", p.unit,
            price, p.stock, round(price * p.stock, 2),
            "Yes" if p.is_active else "No",
        ])
        ws.cell(row=ws.max_row, column=10).number_format = "#,##0"
        ws.cell(row=ws.max_row, column=12).number_format = "#,##0"

    total_value = round(sum(float(p.price) * p.stock for p in products), 2)
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=9, value="TOTAL").font = Font(bold=True)
    tv = ws.cell(row=total_row, column=12, value=total_value)
    tv.font = Font(bold=True)
    tv.number_format = "#,##0"

    widths = [16, 34, 22, 16, 16, 16, 14, 10, 8, 14, 8, 16, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventory_{stamp}.xlsx"'},
    )


def _generate_sku(db: Session) -> str:
    """Auto-assigns the next MM-NNN SKU for products created without one
    (e.g. from the desktop till's inventory tab, which doesn't ask for a
    SKU)."""
    existing = [row[0] for row in db.query(Product.sku).filter(Product.sku.like("MM-%")).all()]
    max_n = 0
    for sku in existing:
        suffix = sku[3:]
        if suffix.isdigit():
            max_n = max(max_n, int(suffix))
    return f"MM-{max_n + 1:03d}"


@router.post("/products/", response_model=ProductAdminOut, status_code=status.HTTP_201_CREATED)
def admin_create_product(
    body: ProductCreateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    data = body.model_dump()
    sku = data.pop("sku") or _generate_sku(db)
    if db.query(Product).filter(Product.sku == sku).first():
        raise HTTPException(status_code=409, detail=f"SKU '{sku}' already exists")
    data["category_zh"] = category_zh_for(data.get("category"))
    product = Product(sku=sku, **data)
    db.add(product)
    db.commit()
    db.refresh(product)
    out = ProductAdminOut.model_validate(product)
    out.low_stock = product.stock <= LOW_STOCK_THRESHOLD
    return out


@router.put("/products/{product_id}", response_model=ProductAdminOut)
def admin_update_product(
    product_id: int,
    body: ProductUpdateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    fields = body.model_dump(exclude_unset=True)
    new_sku = fields.get("sku")
    if new_sku and new_sku != product.sku:
        if db.query(Product).filter(Product.sku == new_sku, Product.id != product_id).first():
            raise HTTPException(status_code=409, detail=f"SKU '{new_sku}' already exists")
    elif "sku" in fields and not new_sku:
        # never allow clearing a SKU (it's required and used as a stable key)
        fields.pop("sku")
    for field, value in fields.items():
        setattr(product, field, value)
    # keep the Chinese category label in sync with the category
    if "category" in fields:
        product.category_zh = category_zh_for(product.category)
    db.commit()
    db.refresh(product)
    out = ProductAdminOut.model_validate(product)
    out.low_stock = product.stock <= LOW_STOCK_THRESHOLD
    return out


@router.post("/products/{product_id}/image", response_model=ProductAdminOut)
def admin_upload_product_image(
    product_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, or WEBP images are accepted")

    contents = file.file.read()
    if len(contents) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be under 5MB")
    try:
        Image.open(io.BytesIO(contents)).verify()
    except UnidentifiedImageError:
        raise HTTPException(status_code=400, detail="File is not a valid image")

    product.image_url = save_product_image(product_id, ext, contents)
    db.commit()
    db.refresh(product)
    out = ProductAdminOut.model_validate(product)
    out.low_stock = product.stock <= LOW_STOCK_THRESHOLD
    return out


# ── Orders ────────────────────────────────────────────────────────────────────

def _build_order_out(order: Order) -> OrderAdminOut:
    return OrderAdminOut(
        id=order.id,
        worker_id=order.worker_id,
        worker_employee_id=order.worker.employee_id,
        worker_name=order.worker.name,
        status=order.status,
        total=float(order.total),
        payment_status=order.payment_status,
        payment_method=order.payment_method,
        created_at=order.created_at,
        items=[
            OrderItemAdminOut(
                product_id=item.product_id,
                product_name=item.product.name if item.product else f"Product #{item.product_id}",
                quantity=item.quantity,
                unit_price=float(item.unit_price),
            )
            for item in order.items
        ],
    )


@router.get("/orders/", response_model=list[OrderAdminOut])
def admin_list_orders(
    status_filter: str | None = Query(None, alias="status"),
    worker_id: int | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    q = db.query(Order).options(
        joinedload(Order.worker),
        joinedload(Order.items).joinedload(OrderItem.product),
    )
    if status_filter:
        q = q.filter(Order.status == status_filter)
    if worker_id:
        q = q.filter(Order.worker_id == worker_id)
    if date_from:
        q = q.filter(Order.created_at >= date_from)
    if date_to:
        q = q.filter(Order.created_at <= date_to)
    orders = q.order_by(Order.created_at.desc()).all()
    return [_build_order_out(o) for o in orders]


@router.put("/orders/{order_id}/ready", response_model=OrderAdminOut)
def mark_order_ready(
    order_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    order = db.query(Order).options(
        joinedload(Order.worker),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != "pending":
        raise HTTPException(status_code=400, detail=f"Order is already '{order.status}'")
    if order.payment_status != "paid":
        raise HTTPException(status_code=400, detail="Order has not been paid yet")
    order.status = "ready"
    order.updated_at = datetime.now(timezone.utc)
    db.commit()
    return _build_order_out(order)


@router.put("/orders/{order_id}/fulfill", response_model=OrderAdminOut)
def fulfill_order(
    order_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    order = db.query(Order).options(
        joinedload(Order.worker),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != "ready":
        raise HTTPException(status_code=400, detail=f"Order must be 'ready' before fulfilling (current: '{order.status}')")
    order.status = "fulfilled"
    order.updated_at = datetime.now(timezone.utc)
    db.commit()
    return _build_order_out(order)


@router.put("/orders/{order_id}/cancel", response_model=OrderAdminOut)
def cancel_order(
    order_id: int,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    order = db.query(Order).options(
        joinedload(Order.worker),
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status not in ("pending", "ready"):
        raise HTTPException(status_code=400, detail=f"Cannot cancel an order with status '{order.status}'")
    # Restore stock
    for item in order.items:
        product = db.get(Product, item.product_id)
        if product:
            product.stock += item.quantity

    wallet_payment = (
        db.query(WalletTransaction)
        .filter(
            WalletTransaction.order_id == order.id,
            WalletTransaction.type == "payment",
        )
        .first()
    )
    if wallet_payment:
        worker = (
            db.query(Worker)
            .filter(Worker.id == order.worker_id)
            .with_for_update()
            .first()
        )
        if worker:
            refund_amount = float(wallet_payment.amount)
            worker.balance = round(float(worker.balance) + refund_amount, 2)
            db.add(WalletTransaction(
                worker_id=worker.id,
                type="refund",
                amount=refund_amount,
                balance_after=worker.balance,
                order_id=order.id,
                performed_by_worker_id=admin.id,
                note="Order cancelled",
            ))

    order.status = "cancelled"
    order.updated_at = datetime.now(timezone.utc)
    db.commit()
    return _build_order_out(order)


@router.post("/orders/{order_id}/refund", response_model=OrderRefundResult)
def refund_order(
    order_id: int,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    """Reverses a cashier-till sale: restores stock, refunds the worker's
    wallet, and logs it. This is the "undo" for a synced till sale — the
    precise inverse of create_cashier_sale. Only wallet-paid orders (i.e.
    cashier-till sales) are eligible; QRIS-paid web pre-orders use a
    different flow (cancel_order above)."""
    order = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.id == order_id)
        .with_for_update()
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status == "cancelled":
        raise HTTPException(status_code=400, detail="Order already refunded")
    if order.payment_method != "wallet":
        raise HTTPException(status_code=400, detail="Only wallet-paid orders can be refunded this way")

    for item in order.items:
        product = (
            db.query(Product)
            .filter(Product.id == item.product_id)
            .with_for_update()
            .first()
        )
        if product:
            product.stock += item.quantity

    worker = (
        db.query(Worker)
        .filter(Worker.id == order.worker_id)
        .with_for_update()
        .first()
    )
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    worker.balance = round(float(worker.balance) + float(order.total), 2)
    db.add(WalletTransaction(
        worker_id=worker.id,
        type="refund",
        amount=float(order.total),
        balance_after=worker.balance,
        order_id=order.id,
        performed_by_worker_id=admin.id,
        note="Cashier sale refund",
    ))

    order.status = "cancelled"
    order.payment_status = "refunded"
    order.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(worker)
    return OrderRefundResult(
        order_id=order.id,
        worker_id=worker.id,
        worker_employee_id=worker.employee_id,
        worker_balance=float(worker.balance),
        refunded_amount=float(order.total),
    )


@router.put("/orders/{order_id}/edit", response_model=OrderAdminOut)
def edit_cashier_sale(
    order_id: int,
    body: OrderEditRequest,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    """In-place correction of a completed cashier (wallet) sale: replace its
    line items, reassign it to another worker, and/or backdate it. Product
    stock and worker wallet balances are reconciled by the difference and the
    change is written to the wallet ledger so everything still adds up.

    Only wallet-paid, non-cancelled orders (i.e. till sales) are editable
    this way; QRIS web pre-orders use cancel/refund instead."""
    if not body.items:
        raise HTTPException(status_code=400, detail="A sale must have at least one item")

    order = (
        db.query(Order)
        .options(joinedload(Order.items).joinedload(OrderItem.product), joinedload(Order.worker))
        .filter(Order.id == order_id)
        .with_for_update()
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.payment_method != "wallet" or order.status == "cancelled":
        raise HTTPException(status_code=400, detail="Only completed cashier (wallet) sales can be edited")

    occurred_at = normalize_occurred_at(body.occurred_at)

    old_total = float(order.total)
    old_worker = (
        db.query(Worker).filter(Worker.id == order.worker_id).with_for_update().first()
    )
    new_worker = old_worker
    if body.worker_id is not None and body.worker_id != order.worker_id:
        new_worker = (
            db.query(Worker).filter(Worker.id == body.worker_id).with_for_update().first()
        )
        if not new_worker or new_worker.role != "worker":
            raise HTTPException(status_code=404, detail="Target worker not found")
        if not new_worker.is_active:
            raise HTTPException(status_code=400, detail="Target worker is inactive")

    # keep the original per-unit prices for products already on the order so a
    # quantity fix doesn't silently reprice history
    old_unit_price = {it.product_id: float(it.unit_price) for it in order.items}

    # 1. restore stock from the old lines, then drop them
    for it in list(order.items):
        product = db.query(Product).filter(Product.id == it.product_id).with_for_update().first()
        if product:
            product.stock += it.quantity
        db.delete(it)
    db.flush()

    # 2. apply the new lines
    new_total = 0.0
    for req in body.items:
        if req.quantity <= 0:
            raise HTTPException(status_code=400, detail="Quantity must be positive")
        product = (
            db.query(Product).filter(Product.id == req.product_id).with_for_update().first()
        )
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {req.product_id} not found")
        if product.stock < req.quantity:
            raise HTTPException(
                status_code=409,
                detail=f"'{product.name}' only has {product.stock} in stock, needs {req.quantity}",
            )
        unit_price = old_unit_price.get(product.id, float(product.price))
        subtotal = round(unit_price * req.quantity, 2)
        new_total += subtotal
        product.stock -= req.quantity
        db.add(OrderItem(
            order_id=order.id, product_id=product.id, quantity=req.quantity,
            unit_price=unit_price, subtotal=subtotal,
        ))
    new_total = round(new_total, 2)

    # 3. reconcile wallet balance(s) + ledger
    if new_worker.id == old_worker.id:
        delta = round(new_total - old_total, 2)
        if delta > 0:
            if float(old_worker.balance) < delta:
                raise HTTPException(status_code=409, detail=f"Insufficient balance for {old_worker.name}")
            old_worker.balance = round(float(old_worker.balance) - delta, 2)
            db.add(WalletTransaction(
                worker_id=old_worker.id, type="adjustment_debit", amount=delta,
                balance_after=old_worker.balance, order_id=order.id,
                performed_by_worker_id=admin.id, note=f"Sale #{order.id} edited",
            ))
        elif delta < 0:
            old_worker.balance = round(float(old_worker.balance) + abs(delta), 2)
            db.add(WalletTransaction(
                worker_id=old_worker.id, type="adjustment_credit", amount=abs(delta),
                balance_after=old_worker.balance, order_id=order.id,
                performed_by_worker_id=admin.id, note=f"Sale #{order.id} edited",
            ))
    else:
        old_worker.balance = round(float(old_worker.balance) + old_total, 2)
        db.add(WalletTransaction(
            worker_id=old_worker.id, type="refund", amount=old_total,
            balance_after=old_worker.balance, order_id=order.id,
            performed_by_worker_id=admin.id, note=f"Sale #{order.id} reassigned away",
        ))
        if float(new_worker.balance) < new_total:
            raise HTTPException(status_code=409, detail=f"Insufficient balance for {new_worker.name}")
        new_worker.balance = round(float(new_worker.balance) - new_total, 2)
        db.add(WalletTransaction(
            worker_id=new_worker.id, type="payment", amount=new_total,
            balance_after=new_worker.balance, order_id=order.id,
            performed_by_worker_id=admin.id, note=f"Sale #{order.id} reassigned here",
        ))

    order.worker_id = new_worker.id
    order.total = new_total
    if occurred_at is not None:
        order.created_at = occurred_at
    order.updated_at = datetime.now(timezone.utc)

    db.commit()

    fresh = (
        db.query(Order)
        .options(joinedload(Order.worker), joinedload(Order.items).joinedload(OrderItem.product))
        .filter(Order.id == order.id)
        .first()
    )
    return _build_order_out(fresh)


# ── Workers ───────────────────────────────────────────────────────────────────

@router.get("/workers/", response_model=list[WorkerOut])
def admin_list_workers(
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    return db.query(Worker).order_by(Worker.employee_id).all()


@router.post("/workers/", response_model=WorkerOut, status_code=status.HTTP_201_CREATED)
def admin_create_worker(
    body: WorkerCreateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    if db.query(Worker).filter(Worker.employee_id == body.employee_id).first():
        raise HTTPException(status_code=409, detail=f"Employee ID '{body.employee_id}' already exists")
    worker = Worker(
        employee_id=body.employee_id,
        hr_employee_id=body.hr_employee_id,
        name=body.name,
        pin_hash=hash_pin(body.pin),
        role="worker",
        is_active=True,
    )
    db.add(worker)
    db.commit()
    db.refresh(worker)
    return worker


@router.patch("/workers/{worker_id}", response_model=WorkerOut)
def admin_update_worker(
    worker_id: int,
    body: WorkerUpdateRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    worker = db.get(Worker, worker_id)
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if body.name is not None:
        worker.name = body.name
    if body.hr_employee_id is not None:
        worker.hr_employee_id = body.hr_employee_id
    if body.pin is not None:
        worker.pin_hash = hash_pin(body.pin)
    if body.is_active is not None:
        worker.is_active = body.is_active
    db.commit()
    db.refresh(worker)
    return worker


@router.post("/workers/{worker_id}/topup", response_model=WalletTopUpResult)
def admin_top_up_worker(
    worker_id: int,
    body: WalletTopUpRequest,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    worker = (
        db.query(Worker)
        .filter(Worker.id == worker_id)
        .with_for_update()
        .first()
    )
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if not worker.is_active:
        raise HTTPException(status_code=400, detail="Cannot top up an inactive worker")

    amount = round(float(body.amount), 2)
    worker.balance = round(float(worker.balance) + amount, 2)
    tx = WalletTransaction(
        worker_id=worker.id,
        type="topup",
        amount=amount,
        balance_after=worker.balance,
        performed_by_worker_id=admin.id,
        note=body.note,
    )
    occurred_at = normalize_occurred_at(body.occurred_at)
    if occurred_at is not None:
        tx.created_at = occurred_at
    db.add(tx)
    db.commit()
    db.refresh(worker)
    db.refresh(tx)
    return WalletTopUpResult(worker=worker, transaction=tx)


@router.post("/workers/{worker_id}/adjust-balance", response_model=WalletTopUpResult)
def admin_adjust_worker_balance(
    worker_id: int,
    body: WalletAdjustRequest,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    """Directly corrects a worker's balance to a specific value (e.g. fixing
    a data-entry mistake), rather than adding a top-up. Still logged as a
    wallet_transaction — type adjustment_credit/adjustment_debit, amount
    stored as the positive size of the change — so the ledger always
    reconciles with the balance."""
    worker = (
        db.query(Worker)
        .filter(Worker.id == worker_id)
        .with_for_update()
        .first()
    )
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    new_balance = round(float(body.new_balance), 2)
    delta = round(new_balance - float(worker.balance), 2)
    if delta == 0:
        raise HTTPException(status_code=400, detail="New balance matches the current balance")

    worker.balance = new_balance
    tx = WalletTransaction(
        worker_id=worker.id,
        type="adjustment_credit" if delta > 0 else "adjustment_debit",
        amount=abs(delta),
        balance_after=worker.balance,
        performed_by_worker_id=admin.id,
        note=body.note,
    )
    occurred_at = normalize_occurred_at(body.occurred_at)
    if occurred_at is not None:
        tx.created_at = occurred_at
    db.add(tx)
    db.commit()
    db.refresh(worker)
    db.refresh(tx)
    return WalletTopUpResult(worker=worker, transaction=tx)


@router.get("/workers/{worker_id}/transactions", response_model=list[WalletTransactionOut])
def admin_worker_transactions(
    worker_id: int,
    db: Session = Depends(get_db),
    _admin: Worker = Depends(require_admin),
):
    worker = db.get(Worker, worker_id)
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    return (
        db.query(WalletTransaction)
        .filter(WalletTransaction.worker_id == worker.id)
        .order_by(WalletTransaction.created_at.desc(), WalletTransaction.id.desc())
        .all()
    )


@router.post("/wallet-transactions/{tx_id}/reverse", response_model=WalletReversalResult)
def reverse_wallet_transaction(
    tx_id: int,
    db: Session = Depends(get_db),
    admin: Worker = Depends(require_admin),
):
    """Reverses a top-up: this is the "undo" for a wallet top-up. Debits
    back the credited amount (deliberately allowed to go negative — this is
    an administrative correction undoing a mistaken credit, not a purchase,
    so there is no floor check here on purpose) and logs a 'reversal'
    transaction. Guarded by tx.reversed so the same top-up can't be
    reversed twice."""
    tx = (
        db.query(WalletTransaction)
        .filter(WalletTransaction.id == tx_id)
        .with_for_update()
        .first()
    )
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    if tx.type != "topup":
        raise HTTPException(status_code=400, detail="Only top-ups can be reversed this way")
    if tx.reversed:
        raise HTTPException(status_code=400, detail="This top-up has already been reversed")

    worker = (
        db.query(Worker)
        .filter(Worker.id == tx.worker_id)
        .with_for_update()
        .first()
    )
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    worker.balance = round(float(worker.balance) - float(tx.amount), 2)
    tx.reversed = True
    db.add(WalletTransaction(
        worker_id=worker.id,
        type="reversal",
        amount=float(tx.amount),
        balance_after=worker.balance,
        order_id=None,
        performed_by_worker_id=admin.id,
        note=f"Reversal of top-up #{tx.id}",
    ))

    db.commit()
    db.refresh(worker)
    return WalletReversalResult(
        transaction_id=tx.id,
        worker_id=worker.id,
        worker_employee_id=worker.employee_id,
        worker_balance=float(worker.balance),
        reversed_amount=float(tx.amount),
    )


# ── Dashboard ─────────────────────────────────────────────────────────────────

def _day_bounds(day: date) -> tuple[datetime, datetime]:
    """[start, end) for a calendar day, matched against naive-UTC created_at
    (the same instant range the rest of the app uses for 'that day')."""
    start = datetime.combine(day, time.min)
    return start, start + timedelta(days=1)


@router.get("/dashboard/summary")
def dashboard_summary(
    day: date | None = Query(None, description="calendar day, defaults to today (UTC)"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    """Single-day totals for the admin dashboard: money spent in the shop
    (belanja) vs money paid into wallets (deposit), kept separate and scoped
    to one day rather than an all-time running total."""
    target = day or datetime.utcnow().date()
    start, end = _day_bounds(target)

    orders = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.created_at >= start, Order.created_at < end, Order.status != "cancelled")
        .all()
    )
    sales_total = round(sum(float(o.total) for o in orders), 2)
    items_sold = sum(it.quantity for o in orders for it in o.items)

    topups = (
        db.query(WalletTransaction)
        .filter(
            WalletTransaction.created_at >= start,
            WalletTransaction.created_at < end,
            WalletTransaction.type == "topup",
        )
        .all()
    )
    deposits_total = round(sum(float(t.amount) for t in topups), 2)

    return {
        "day": target.isoformat(),
        "sales_total": sales_total,
        "deposits_total": deposits_total,
        "order_count": len(orders),
        "items_sold": items_sold,
    }


# ── Reports ───────────────────────────────────────────────────────────────────

def _spending_query(db: Session, date_from: datetime | None, date_to: datetime | None):
    q = (
        db.query(Worker, Order)
        .join(Order, Order.worker_id == Worker.id)
        .filter(Order.status != "cancelled")
    )
    if date_from:
        q = q.filter(Order.created_at >= date_from)
    if date_to:
        q = q.filter(Order.created_at <= date_to)
    return q.all()


@router.get("/reports/spending", response_model=list[WorkerSpending])
def spending_report(
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    rows = _spending_query(db, date_from, date_to)
    totals: dict[int, WorkerSpending] = {}
    for worker, order in rows:
        if worker.id not in totals:
            totals[worker.id] = WorkerSpending(
                employee_id=worker.employee_id,
                hr_employee_id=worker.hr_employee_id,
                name=worker.name,
                total_deduction=0.0,
                order_count=0,
            )
        totals[worker.id].total_deduction += float(order.total)
        totals[worker.id].order_count += 1
    for s in totals.values():
        s.total_deduction = round(s.total_deduction, 2)
    return sorted(totals.values(), key=lambda s: s.employee_id)


@router.get("/reports/spending.csv")
def spending_report_csv(
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    month: str | None = Query(None, description="e.g. 2026-03"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    # Allow ?month=2026-03 as a shortcut
    if month and not date_from and not date_to:
        try:
            year, mon = map(int, month.split("-"))
            date_from = datetime(year, mon, 1, tzinfo=timezone.utc)
            next_mon = mon % 12 + 1
            next_year = year + (1 if mon == 12 else 0)
            date_to = datetime(next_year, next_mon, 1, tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format, use YYYY-MM")

    rows = _spending_query(db, date_from, date_to)
    totals: dict[int, dict] = {}
    for worker, order in rows:
        if worker.id not in totals:
            totals[worker.id] = {
                "worker_name": worker.name,
                "employee_id": worker.employee_id,
                "hr_employee_id": worker.hr_employee_id or "",
                "total_deduction": 0.0,
                "month": month or (date_from.strftime("%Y-%m") if date_from else "all"),
            }
        totals[worker.id]["total_deduction"] += float(order.total)

    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=["worker_name", "employee_id", "hr_employee_id", "total_deduction", "month"],
    )
    writer.writeheader()
    for row in sorted(totals.values(), key=lambda r: r["employee_id"]):
        row["total_deduction"] = round(row["total_deduction"], 2)
        writer.writerow(row)

    buf.seek(0)
    filename = f"spending_{month or 'report'}.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


_ORDER_TYPE_LABELS = {"qris": "QRIS pre-order", "wallet": "Cashier sale"}
_WTX_TYPE_LABELS = {
    "topup": "Top-up",
    "payment": "Payment",
    "refund": "Refund",
    "reversal": "Reversal",
    "adjustment_credit": "Adjustment (+)",
    "adjustment_debit": "Adjustment (−)",
}
# types that move a worker's balance down — shown as a negative amount
_WTX_NEGATIVE = {"payment", "reversal", "adjustment_debit"}

_MONEY_FMT = "#,##0"


def _daily_reconciliation(orders, wtxs) -> dict[str, dict]:
    """Per-day rollup: deposits (wallet top-ups) vs shop sales (active order
    totals), keyed by 'YYYY-MM-DD'."""
    from collections import defaultdict

    days: dict[str, dict] = defaultdict(lambda: {"deposits": 0.0, "sales": 0.0, "orders": 0})
    seen_orders: set[int] = set()
    for o in orders:
        if o.status == "cancelled" or o.id in seen_orders:
            continue
        seen_orders.add(o.id)
        d = o.created_at.strftime("%Y-%m-%d")
        days[d]["sales"] += float(o.total)
        days[d]["orders"] += 1
    for tx in wtxs:
        if tx.type != "topup":
            continue
        d = tx.created_at.strftime("%Y-%m-%d")
        days[d]["deposits"] += float(tx.amount)
    return days


def _write_daily_sheet(ws, day_rows: list[tuple[str, dict]]) -> None:
    """day_rows: list of (date_str, {deposits, sales, orders}) already ordered."""
    ws.append(["Date", "Deposits (Rp)", "Sales (Rp)", "Orders", "Net (Rp)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    tot_dep = tot_sales = 0.0
    tot_orders = 0
    for d, v in day_rows:
        dep, sales, orders = round(v["deposits"], 2), round(v["sales"], 2), v["orders"]
        tot_dep += dep
        tot_sales += sales
        tot_orders += orders
        ws.append([d, dep, sales, orders, round(dep - sales, 2)])
        for col in (2, 3, 5):
            ws.cell(row=ws.max_row, column=col).number_format = _MONEY_FMT

    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for col, val in ((2, round(tot_dep, 2)), (3, round(tot_sales, 2)),
                     (4, tot_orders), (5, round(tot_dep - tot_sales, 2))):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(bold=True)
        if col != 4:
            c.number_format = _MONEY_FMT

    for i, w in enumerate([14, 16, 16, 9, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


@router.get("/reports/transactions.xlsx")
def transactions_export_xlsx(
    date_from: date | None = Query(None, description="start date, inclusive"),
    date_to: date | None = Query(None, description="end date, inclusive"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    """Every sale (broken out to one row per line item) and every wallet
    movement in the date range, as a two-sheet .xlsx workbook for daily
    bookkeeping.

    created_at is stored (and shown everywhere else in the app) as naive
    UTC, so date_from/date_to are matched against it directly — a given
    calendar day here is the same instant range the order list, dashboard
    and spending report use for that day."""
    start = datetime.combine(date_from, time.min) if date_from else None
    end = datetime.combine(date_to, time.min) + timedelta(days=1) if date_to else None

    order_q = db.query(Order).options(
        joinedload(Order.worker),
        joinedload(Order.items).joinedload(OrderItem.product),
    )
    wtx_q = db.query(WalletTransaction).options(
        joinedload(WalletTransaction.worker),
        joinedload(WalletTransaction.performed_by),
    )
    if start is not None:
        order_q = order_q.filter(Order.created_at >= start)
        wtx_q = wtx_q.filter(WalletTransaction.created_at >= start)
    if end is not None:
        order_q = order_q.filter(Order.created_at < end)
        wtx_q = wtx_q.filter(WalletTransaction.created_at < end)

    orders = order_q.order_by(Order.created_at, Order.id).all()
    wtxs = wtx_q.order_by(WalletTransaction.created_at, WalletTransaction.id).all()

    wb = Workbook()

    # ── Sheet 1: Sales — one row per order line item ─────────────────────────
    ws = wb.active
    ws.title = "Sales"
    ws.append([
        "Date", "Time", "Order #", "Type", "Status", "Payment",
        "Employee ID", "Worker", "SKU", "Product",
        "Qty", "Unit price (Rp)", "Line total (Rp)",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    sales_total = 0.0
    sales_total_active = 0.0  # excludes cancelled orders
    for o in orders:
        otype = _ORDER_TYPE_LABELS.get(o.payment_method or "", o.payment_method or "—")
        for it in sorted(o.items, key=lambda x: x.product.name if x.product else ""):
            line_total = float(it.subtotal)
            sales_total += line_total
            if o.status != "cancelled":
                sales_total_active += line_total
            ws.append([
                o.created_at.strftime("%Y-%m-%d"),
                o.created_at.strftime("%H:%M"),
                o.id, otype, o.status, o.payment_status,
                o.worker.employee_id if o.worker else "",
                o.worker.name if o.worker else "",
                it.product.sku if it.product else "",
                it.product.name if it.product else f"Product #{it.product_id}",
                it.quantity, float(it.unit_price), line_total,
            ])
            ws.cell(row=ws.max_row, column=12).number_format = _MONEY_FMT
            ws.cell(row=ws.max_row, column=13).number_format = _MONEY_FMT

    trow = ws.max_row + 2
    ws.cell(row=trow, column=10, value="TOTAL").font = Font(bold=True)
    tc = ws.cell(row=trow, column=13, value=round(sales_total, 2))
    tc.font, tc.number_format = Font(bold=True), _MONEY_FMT
    ws.cell(row=trow + 1, column=10, value="TOTAL (excl. cancelled)").font = Font(bold=True)
    tc = ws.cell(row=trow + 1, column=13, value=round(sales_total_active, 2))
    tc.font, tc.number_format = Font(bold=True), _MONEY_FMT

    for i, w in enumerate([12, 7, 8, 15, 11, 10, 13, 20, 12, 30, 6, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Wallet movements — one row per wallet_transaction ────────────
    ws2 = wb.create_sheet("Wallet movements")
    ws2.append([
        "Date", "Time", "Type", "Employee ID", "Worker",
        "Amount (Rp)", "Balance after (Rp)", "Order #", "Performed by", "Note",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"

    wtx_net = 0.0
    for tx in wtxs:
        signed = -float(tx.amount) if tx.type in _WTX_NEGATIVE else float(tx.amount)
        wtx_net += signed
        ws2.append([
            tx.created_at.strftime("%Y-%m-%d"),
            tx.created_at.strftime("%H:%M"),
            _WTX_TYPE_LABELS.get(tx.type, tx.type),
            tx.worker.employee_id if tx.worker else "",
            tx.worker.name if tx.worker else "",
            round(signed, 2), float(tx.balance_after),
            tx.order_id or "",
            tx.performed_by.name if tx.performed_by else "",
            tx.note or "",
        ])
        ws2.cell(row=ws2.max_row, column=6).number_format = _MONEY_FMT
        ws2.cell(row=ws2.max_row, column=7).number_format = _MONEY_FMT

    trow2 = ws2.max_row + 2
    ws2.cell(row=trow2, column=5, value="NET MOVEMENT").font = Font(bold=True)
    nc = ws2.cell(row=trow2, column=6, value=round(wtx_net, 2))
    nc.font, nc.number_format = Font(bold=True), _MONEY_FMT

    for i, w in enumerate([12, 7, 16, 13, 20, 15, 16, 9, 18, 32], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Daily summary — one row per day (deposits vs sales) ──────────
    ws3 = wb.create_sheet("Daily summary")
    days = _daily_reconciliation(orders, wtxs)
    _write_daily_sheet(ws3, sorted(days.items()))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    span = f"{date_from}_{date_to}" if date_from and date_to else "all"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="transactions_{span}.xlsx"'},
    )


@router.get("/reports/monthly.xlsx")
def monthly_reconciliation_xlsx(
    month: str = Query(..., description="e.g. 2026-08"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    """One workbook for a whole month's bookkeeping: a 'Daily' sheet with a
    row for every calendar day (deposits vs shop sales, and the net), plus a
    'By worker' sheet totalling each worker's shop spend for the month (the
    figure that gets deducted from payroll)."""
    try:
        year, mon = map(int, month.split("-"))
        start = datetime(year, mon, 1)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid month, use YYYY-MM")
    end = datetime(year + (mon == 12), (mon % 12) + 1, 1)

    orders = (
        db.query(Order)
        .options(joinedload(Order.worker))
        .filter(Order.created_at >= start, Order.created_at < end, Order.status != "cancelled")
        .all()
    )
    topups = (
        db.query(WalletTransaction)
        .filter(
            WalletTransaction.created_at >= start,
            WalletTransaction.created_at < end,
            WalletTransaction.type == "topup",
        )
        .all()
    )

    days = _daily_reconciliation(orders, topups)
    # fill in every calendar day, even quiet ones, so the month reads top-to-bottom
    ordered: list[tuple[str, dict]] = []
    cursor = start
    while cursor < end:
        key = cursor.strftime("%Y-%m-%d")
        ordered.append((key, days.get(key, {"deposits": 0.0, "sales": 0.0, "orders": 0})))
        cursor += timedelta(days=1)

    wb = Workbook()
    _write_daily_sheet(wb.active, ordered)
    wb.active.title = "Daily"

    ws2 = wb.create_sheet("By worker")
    ws2.append(["Employee ID", "HR / payroll ID", "Worker", "Orders", "Shop spend (Rp)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"

    by_worker: dict[int, dict] = {}
    for o in orders:
        w = o.worker
        rec = by_worker.setdefault(o.worker_id, {
            "employee_id": w.employee_id if w else "",
            "hr": (w.hr_employee_id or "") if w else "",
            "name": w.name if w else "",
            "orders": 0, "spend": 0.0,
        })
        rec["orders"] += 1
        rec["spend"] += float(o.total)

    grand = 0.0
    for rec in sorted(by_worker.values(), key=lambda r: r["employee_id"]):
        spend = round(rec["spend"], 2)
        grand += spend
        ws2.append([rec["employee_id"], rec["hr"], rec["name"], rec["orders"], spend])
        ws2.cell(row=ws2.max_row, column=5).number_format = _MONEY_FMT

    r = ws2.max_row + 2
    ws2.cell(row=r, column=3, value="TOTAL").font = Font(bold=True)
    tc = ws2.cell(row=r, column=5, value=round(grand, 2))
    tc.font, tc.number_format = Font(bold=True), _MONEY_FMT
    for i, w in enumerate([13, 18, 22, 9, 18], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="reconciliation_{month}.xlsx"'},
    )
