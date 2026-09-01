"""Module 05 — Admin Dashboard API tests."""

import io

from openpyxl import load_workbook

from app.models.order import Order, OrderItem
from app.models.wallet import WalletTransaction


def _place_order(client, worker_token, product_id, quantity):
    return client.post(
        "/api/orders/",
        headers={"Authorization": f"Bearer {worker_token}"},
        json={"items": [{"product_id": product_id, "quantity": quantity}]},
    ).json()


def _mark_order_ready(db_session, order_id):
    order = db_session.get(Order, order_id)
    order.payment_status = "paid"
    order.status = "ready"
    db_session.commit()


def test_admin_list_products_shows_low_stock(client, admin_token, products):
    client.post(
        "/api/admin/products/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"name": "Bulk Water", "sku": "BW-001", "price": 1.0, "stock": 100},
    )
    resp = client.get("/api/admin/products/", headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    by_sku = {p["sku"]: p for p in resp.json()}
    assert by_sku["SH-001"]["low_stock"] is True    # stock=5, threshold=30
    assert by_sku["WG-001"]["low_stock"] is True    # stock=10, threshold=30
    assert by_sku["BW-001"]["low_stock"] is False   # stock=100


def test_admin_create_product(client, admin_token):
    resp = client.post(
        "/api/admin/products/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"name": "Protein Bar", "sku": "PB-001", "price": 2.50, "stock": 3},
    )
    assert resp.status_code == 201
    data = resp.json()
    assert data["name"] == "Protein Bar"
    assert data["low_stock"] is True  # stock=3 <= 30


def test_admin_update_product_sku_and_conflict(client, admin_token, products):
    p0, p1 = products[0].id, products[1].id
    ok = client.put(
        f"/api/admin/products/{p0}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"sku": "WG-999"},
    )
    assert ok.status_code == 200
    assert ok.json()["sku"] == "WG-999"

    clash = client.put(
        f"/api/admin/products/{p1}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"sku": "WG-999"},
    )
    assert clash.status_code == 409


def test_admin_create_product_duplicate_sku(client, admin_token, products):
    resp = client.post(
        "/api/admin/products/",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"name": "Dupe", "sku": "WG-001", "price": 1.0, "stock": 1},
    )
    assert resp.status_code == 409


def test_admin_update_product(client, admin_token, products):
    product_id = products[0].id
    resp = client.put(
        f"/api/admin/products/{product_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"price": 9.99, "stock": 50},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["price"] == 9.99
    assert data["stock"] == 50


def test_admin_list_orders(client, admin_token, worker_token, products, worker):
    _place_order(client, worker_token, products[0].id, 1)
    resp = client.get("/api/admin/orders/", headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    orders = resp.json()
    assert len(orders) >= 1
    assert orders[0]["worker_employee_id"] == "W001"


def test_admin_filter_orders_by_status(client, admin_token, worker_token, products):
    order = _place_order(client, worker_token, products[0].id, 1)
    # Fulfill it
    client.put(f"/api/admin/orders/{order['id']}/fulfill",
               headers={"Authorization": f"Bearer {admin_token}"})

    resp = client.get("/api/admin/orders/?status=fulfilled",
                      headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    assert all(o["status"] == "fulfilled" for o in resp.json())


def test_fulfill_order(client, admin_token, worker_token, products, db_session):
    order = _place_order(client, worker_token, products[0].id, 1)
    _mark_order_ready(db_session, order["id"])
    resp = client.put(
        f"/api/admin/orders/{order['id']}/fulfill",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "fulfilled"


def test_cancel_order_restores_stock(client, admin_token, worker_token, products, db_session):
    gloves = products[0]
    initial_stock = gloves.stock
    order = _place_order(client, worker_token, gloves.id, 3)

    resp = client.put(
        f"/api/admin/orders/{order['id']}/cancel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "cancelled"
    db_session.refresh(gloves)
    assert gloves.stock == initial_stock  # stock fully restored


def test_cannot_cancel_fulfilled_order(client, admin_token, worker_token, products, db_session):
    order = _place_order(client, worker_token, products[0].id, 1)
    _mark_order_ready(db_session, order["id"])
    client.put(f"/api/admin/orders/{order['id']}/fulfill",
               headers={"Authorization": f"Bearer {admin_token}"})
    resp = client.put(
        f"/api/admin/orders/{order['id']}/cancel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 400


def test_spending_report(client, admin_token, worker_token, products):
    _place_order(client, worker_token, products[0].id, 2)
    resp = client.get("/api/admin/reports/spending",
                      headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    report = resp.json()
    assert len(report) >= 1
    w = next(r for r in report if r["employee_id"] == "W001")
    assert w["total_deduction"] == round(5.50 * 2, 2)


def test_spending_csv_export(client, admin_token, worker_token, products):
    _place_order(client, worker_token, products[0].id, 1)
    resp = client.get("/api/admin/reports/spending.csv",
                      headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    lines = resp.text.strip().splitlines()
    assert lines[0] == "worker_name,employee_id,total_deduction,month"
    assert "W001" in lines[1]


def _seed_cashier_sale(db_session, worker, product, quantity, created_at=None):
    """A settled wallet sale + its ledger entry, written straight to the DB
    (the HTTP order path needs a pickup slot; a cashier sale doesn't)."""
    unit_price = float(product.price)
    order = Order(
        worker_id=worker.id, status="fulfilled",
        payment_status="paid", payment_method="wallet",
        total=unit_price * quantity,
    )
    if created_at is not None:
        order.created_at = created_at
    db_session.add(order)
    db_session.flush()
    db_session.add(OrderItem(
        order_id=order.id, product_id=product.id, quantity=quantity,
        unit_price=unit_price, subtotal=unit_price * quantity,
    ))
    tx = WalletTransaction(
        worker_id=worker.id, type="payment", amount=unit_price * quantity,
        balance_after=0.0, order_id=order.id, performed_by_worker_id=worker.id,
        note="Cashier sale",
    )
    if created_at is not None:
        tx.created_at = created_at
    db_session.add(tx)
    db_session.commit()
    return order


def test_transactions_xlsx_export(client, admin_token, db_session, worker, products):
    _seed_cashier_sale(db_session, worker, products[0], 2)  # Work Gloves @ 5.50
    resp = client.get(
        "/api/admin/reports/transactions.xlsx",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.headers["content-disposition"].endswith('transactions_all.xlsx"')

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Sales", "Wallet movements", "Daily summary"]

    sales = wb["Sales"]
    assert [c.value for c in sales[1]][:4] == ["Date", "Time", "Order #", "Type"]
    row = [c.value for c in sales[2]]
    assert row[3] == "Cashier sale"      # Type
    assert row[6] == "W001"              # Employee ID
    assert row[9] == "Work Gloves"       # Product
    assert row[10] == 2                  # Qty
    assert row[12] == 11.0              # Line total (2 x 5.50)

    wallet = wb["Wallet movements"]
    wrow = [c.value for c in wallet[2]]
    assert wrow[2] == "Payment"
    assert wrow[5] == -11.0             # payment shown as a negative amount


def test_transactions_xlsx_date_range_filters(client, admin_token, db_session, worker, products):
    _seed_cashier_sale(db_session, worker, products[0], 1)
    # a window entirely in the past should contain no sale rows
    resp = client.get(
        "/api/admin/reports/transactions.xlsx",
        params={"date_from": "2020-01-01", "date_to": "2020-01-31"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-disposition"].endswith('transactions_2020-01-01_2020-01-31.xlsx"')
    sales = load_workbook(io.BytesIO(resp.content))["Sales"]
    data_rows = [r for r in sales.iter_rows(min_row=2, values_only=True) if r[0]]
    assert data_rows == []


def test_transactions_xlsx_single_day_filter(client, admin_token, db_session, worker, products):
    """date_from == date_to matches exactly that calendar day (naive-UTC, the
    same instant range the rest of the app uses for that day) — including a
    sale at 23:00, which must not spill into the next day."""
    from datetime import datetime

    _seed_cashier_sale(db_session, worker, products[0], 1,
                       created_at=datetime(2026, 3, 14, 23, 0))   # late on the 14th
    _seed_cashier_sale(db_session, worker, products[1], 1,
                       created_at=datetime(2026, 3, 15, 8, 30))   # the 15th
    _seed_cashier_sale(db_session, worker, products[0], 1,
                       created_at=datetime(2026, 3, 16, 0, 30))   # early on the 16th

    resp = client.get(
        "/api/admin/reports/transactions.xlsx",
        params={"date_from": "2026-03-15", "date_to": "2026-03-15"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    sales = load_workbook(io.BytesIO(resp.content))["Sales"]
    rows = [r for r in sales.iter_rows(min_row=2, values_only=True) if r[0] and r[2]]
    assert [(r[0], r[1]) for r in rows] == [("2026-03-15", "08:30")]


def test_transactions_xlsx_rejects_worker(client, worker_token):
    resp = client.get(
        "/api/admin/reports/transactions.xlsx",
        headers={"Authorization": f"Bearer {worker_token}"},
    )
    assert resp.status_code == 403


def test_admin_routes_reject_worker(client, worker_token):
    resp = client.get("/api/admin/products/",
                      headers={"Authorization": f"Bearer {worker_token}"})
    assert resp.status_code == 403


# ── Dashboard summary (item 7) ───────────────────────────────────────────────

def test_dashboard_summary_splits_sales_and_deposits(client, admin_token, db_session, worker, products):
    _seed_cashier_sale(db_session, worker, products[0], 2)  # 2 x 5.50 = 11 today
    client.post(
        f"/api/admin/workers/{worker.id}/topup",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"amount": 250},
    )
    resp = client.get("/api/admin/dashboard/summary",
                      headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200
    data = resp.json()
    assert data["sales_total"] == 11.0
    assert data["deposits_total"] == 250.0
    assert data["order_count"] == 1
    assert data["items_sold"] == 2


def test_dashboard_summary_scoped_to_a_single_day(client, admin_token, db_session, worker, products):
    from datetime import datetime
    _seed_cashier_sale(db_session, worker, products[0], 1,
                       created_at=datetime(2026, 3, 15, 9, 0))
    _seed_cashier_sale(db_session, worker, products[1], 1)  # today

    old = client.get("/api/admin/dashboard/summary",
                     params={"day": "2026-03-15"},
                     headers={"Authorization": f"Bearer {admin_token}"}).json()
    assert old["sales_total"] == 5.5
    assert old["order_count"] == 1

    other = client.get("/api/admin/dashboard/summary",
                       params={"day": "2026-03-16"},
                       headers={"Authorization": f"Bearer {admin_token}"}).json()
    assert other["sales_total"] == 0
    assert other["order_count"] == 0


# ── Edit a completed cashier sale (item 4) ───────────────────────────────────

def test_edit_cashier_sale_reconciles_stock_and_wallet(client, admin_token, db_session, worker, products):
    gloves = products[0]  # stock 10, price 5.50
    client.post(
        f"/api/admin/workers/{worker.id}/topup",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"amount": 100},
    )
    db_session.refresh(worker)

    # a settled 2-unit sale, balance debited by hand to mirror the real path
    order = _seed_cashier_sale(db_session, worker, gloves, 2)
    worker.balance = round(float(worker.balance) - 11.0, 2)
    db_session.commit()
    db_session.refresh(gloves)
    stock_after_sale = gloves.stock  # 8

    resp = client.put(
        f"/api/admin/orders/{order.id}/edit",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"items": [{"product_id": gloves.id, "quantity": 1}]},
    )
    assert resp.status_code == 200
    assert resp.json()["total"] == 5.5

    db_session.refresh(gloves)
    db_session.refresh(worker)
    assert gloves.stock == stock_after_sale + 1          # one unit returned
    assert float(worker.balance) == round(100 - 5.5, 2)   # refunded the 5.50 difference
    credit = (
        db_session.query(WalletTransaction)
        .filter(WalletTransaction.type == "adjustment_credit",
                WalletTransaction.order_id == order.id)
        .one()
    )
    assert float(credit.amount) == 5.5


def test_edit_rejects_non_wallet_order(client, admin_token, db_session, worker, products):
    order = Order(worker_id=worker.id, status="pending", payment_status="paid",
                  payment_method="qris", total=5.5)
    db_session.add(order)
    db_session.commit()
    resp = client.put(
        f"/api/admin/orders/{order.id}/edit",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"items": [{"product_id": products[0].id, "quantity": 1}]},
    )
    assert resp.status_code == 400


# ── Monthly reconciliation export (item 5) ───────────────────────────────────

def test_monthly_reconciliation_xlsx(client, admin_token, db_session, worker, products):
    from datetime import datetime
    _seed_cashier_sale(db_session, worker, products[0], 2,
                       created_at=datetime(2026, 3, 10, 12, 0))
    resp = client.get(
        "/api/admin/reports/monthly.xlsx",
        params={"month": "2026-03"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 200
    assert resp.headers["content-disposition"].endswith('reconciliation_2026-03.xlsx"')
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Daily", "By worker"]
    daily = wb["Daily"]
    assert [c.value for c in daily[1]] == ["Date", "Deposits (Rp)", "Sales (Rp)", "Orders", "Net (Rp)"]
    rows = {r[0]: r for r in daily.iter_rows(min_row=2, values_only=True) if r[0] and str(r[0]).startswith("2026")}
    assert rows["2026-03-10"][2] == 11.0   # sales that day
    assert len(rows) == 31                  # every day of March present


def test_backdated_topup_lands_on_chosen_day(client, admin_token, db_session, worker):
    resp = client.post(
        f"/api/admin/workers/{worker.id}/topup",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"amount": 500, "occurred_at": "2026-02-01T08:00:00"},
    )
    assert resp.status_code == 200
    tx = db_session.query(WalletTransaction).filter(WalletTransaction.type == "topup").one()
    assert tx.created_at.strftime("%Y-%m-%d") == "2026-02-01"


def test_backdated_topup_rejects_future(client, admin_token, worker):
    resp = client.post(
        f"/api/admin/workers/{worker.id}/topup",
        headers={"Authorization": f"Bearer {admin_token}"},
        json={"amount": 500, "occurred_at": "2099-01-01T00:00:00"},
    )
    assert resp.status_code == 400
